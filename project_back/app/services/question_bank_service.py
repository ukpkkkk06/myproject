from openpyxl import load_workbook
from fastapi import HTTPException
from app.models.question import Question
from app.models.question_version import QuestionVersion
from app.models.tag import Tag, QuestionTag
from sqlalchemy.orm import Session
from app.schemas.question_bank import ImportQuestionsResult, ImportErrorItem
from typing import Dict
from sqlalchemy import select, exists
import json
from app.models.user import User  # 修复未定义 User

HEADER_EXPECT = ["题干","选项A","选项B","选项C","选项D","题型(单选/多选/填空)","正确答案（单选多选请填入ABCD,填空直接填入答案，不同方式用;隔开如:BEIJNG;beijng）","解析","学科（数学，英语，化学，物理，语文）","学段（小学，初中，高中，大学）"]
ANSWER_KEYS = ["A","B","C","D"]
QUESTION_TYPES = {"单选": "SC", "多选": "MC", "填空": "FILL"}  # 🆕 添加填空题型

def _get_or_none(tag_map: Dict[str, Tag], name: str):
    if not name:
        return None
    return tag_map.get(name.strip())

def import_questions_from_excel(db: Session, file_path: str, user_id: int) -> ImportQuestionsResult:
    try:
        # 🚀 优化：使用只读模式和data_only模式，大幅减少内存占用
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取Excel: {e}")
    ws = wb.active
    header = [ (ws.cell(row=1,column=i+1).value or "").strip() for i in range(len(HEADER_EXPECT)) ]
    if header != HEADER_EXPECT:
        raise HTTPException(status_code=400, detail="模板表头不匹配，请下载最新模板")

    tags = db.execute(select(Tag).where(Tag.type.in_(["SUBJECT","LEVEL"]))).scalars().all()
    tag_map = {t.name.strip(): t for t in tags}

    result = ImportQuestionsResult(total_rows=0, success=0, failed=0, errors=[])

    def cell_str(row:int, col:int) -> str:
        return str(ws.cell(row=row, column=col).value or "").strip()

    for r in range(2, ws.max_row+1):
        # 跳过纯空行
        if all((cell_str(r, c) == "") for c in range(1, 6)):
            continue
        result.total_rows += 1
        try:
            stem = cell_str(r, 1)
            if not stem:
                raise ValueError("题干为空")
            
            # 🔥 检查题干是否重复（仅当前用户、仅激活题目）
            # 明确指定 JOIN 条件，避免歧义
            existing = db.query(QuestionVersion).join(
                Question, 
                QuestionVersion.question_id == Question.id
            ).filter(
                QuestionVersion.stem == stem,
                QuestionVersion.created_by == user_id,
                Question.is_active == True,
                QuestionVersion.is_active == 1
            ).first()
            
            if existing:
                raise ValueError(f"题目重复：您已创建过相同题干的题目（题目ID: {existing.question_id}）")
            
            A = cell_str(r, 2)
            B = cell_str(r, 3)
            C = cell_str(r, 4)
            D = cell_str(r, 5)
            qtype_str = cell_str(r, 6)  # 🆕 题型列（单选/多选）
            correct = cell_str(r, 7).upper()  # 🆕 正确答案移到第7列
            analysis = cell_str(r, 8)  # 🆕 解析移到第8列
            subject_name = cell_str(r, 9)  # 🆕 学科移到第9列
            level_name = cell_str(r, 10)  # 🆕 学段移到第10列
            
            # 🆕 验证题型
            if qtype_str not in QUESTION_TYPES:
                raise ValueError(f"题型必须是'单选'、'多选'或'填空'，当前值：{qtype_str}")
            
            qtype = QUESTION_TYPES[qtype_str]  # SC 或 MC 或 FILL
            
            # 🆕 根据题型验证答案
            if qtype == "SC":
                if not all([A, B, C, D]):
                    raise ValueError("单选题必须填写所有选项A/B/C/D")
                if correct not in ANSWER_KEYS:
                    raise ValueError("单选题正确选项必须是 A/B/C/D 之一")
            elif qtype == "MC":
                if not all([A, B, C, D]):
                    raise ValueError("多选题必须填写所有选项A/B/C/D")
                if not correct or len(correct) < 2:
                    raise ValueError("多选题至少要有2个正确答案")
                if not all(c in ANSWER_KEYS for c in correct):
                    raise ValueError(f"多选题正确选项必须是 A/B/C/D 的组合，如 ABC，当前值：{correct}")
                # 标准化多选答案：去重并排序（例如 "BCA" -> "ABC"）
                correct = "".join(sorted(set(correct)))
            elif qtype == "FILL":
                # 🆕 填空题验证
                if not correct:
                    raise ValueError("填空题答案不能为空，请在'正确答案'列填写文本答案（支持用分号分隔多个答案，如：北京;beijing）")
                # 🆕 提示用户：填空题不需要填写选项
                if any([A, B, C, D]):
                    raise ValueError("填空题不需要填写选项A/B/C/D，请将这些列留空")

            # 🆕 根据题型设置选项
            if qtype == "FILL":
                # 填空题不需要选项
                options = None
            else:
                # 单选题和多选题需要选项
                options = [
                    {"key":"A","text":A},
                    {"key":"B","text":B},
                    {"key":"C","text":C},
                    {"key":"D","text":D},
                ]

            # 🆕 根据题型创建 Question
            q = Question(type=qtype, is_active=True)
            if hasattr(q, "created_by"):
                setattr(q, "created_by", user_id)
            db.add(q)
            db.flush()  # 拿到 q.id

            # 关键：设置 version_no=1，并置 is_active
            qv = QuestionVersion(question_id=q.id, version_no=1, is_active=1)
            setattr(qv, "stem", stem)
            
            # 🆕 根据题型设置 options
            if qtype == "FILL":
                # 填空题不设置 options
                if hasattr(qv, "options"):
                    qv.options = None
            else:
                # 单选题和多选题设置 options
                if hasattr(qv, "options"):
                    qv.options = options
                elif hasattr(qv, "choices"):
                    qv.choices = json.dumps([o["text"] for o in options], ensure_ascii=False)

            if hasattr(qv, "analysis"):
                qv.analysis = analysis
            elif hasattr(qv, "explanation"):
                qv.explanation = analysis

            if hasattr(qv, "correct_answer"):
                qv.correct_answer = correct
            elif hasattr(qv, "answer"):
                qv.answer = correct

            if hasattr(qv, "created_by"):
                qv.created_by = user_id

            db.add(qv)
            db.flush()  # 拿到 qv.id

            if hasattr(q, "current_version_id"):
                q.current_version_id = qv.id
                db.add(q)

            # 关联标签
            subj_tag = _get_or_none(tag_map, subject_name)
            level_tag = _get_or_none(tag_map, level_name)
            if subj_tag:
                db.add(QuestionTag(question_id=q.id, tag_id=subj_tag.id))
            if level_tag:
                db.add(QuestionTag(question_id=q.id, tag_id=level_tag.id))

            # 每行成功后提交
            db.commit()
            result.success += 1

        except Exception as e:
            # 本行失败回滚并记录
            db.rollback()
            result.failed += 1
            result.errors.append(ImportErrorItem(row=r, reason=str(e)))

    # 末尾不再统一 commit
    return result

def list_my_questions(
    db: Session,
    user: User,
    page: int = 1,
    size: int = 10,
    keyword: str | None = None,
    qtype: str | None = None,
    difficulty: int | None = None,
    active_only: bool = False,
    subject_id: int | None = None,
    level_id: int | None = None,
):
    page = max(1, int(page or 1))
    size = max(1, min(int(size or 10), 100))

    q = (
        db.query(
            Question.id.label("question_id"),
            Question.type,
            Question.difficulty,
            Question.audit_status,
            Question.is_active,
            Question.created_at,
            Question.updated_at,
            QuestionVersion.stem,
        )
        .join(QuestionVersion, Question.current_version_id == QuestionVersion.id)
        .filter(QuestionVersion.created_by == user.id)
    )

    if keyword:
        kw = f"%{keyword.strip()}%"
        q = q.filter(QuestionVersion.stem.like(kw))
    if qtype:
        q = q.filter(Question.type == qtype)
    if difficulty is not None:
        q = q.filter(Question.difficulty == difficulty)
    if active_only:
        q = q.filter(Question.is_active == True)
    if subject_id:
        q = q.filter(
            exists().where(
                (QuestionTag.question_id == Question.id) &
                (QuestionTag.tag_id == subject_id)
            )
        )
    if level_id:
        q = q.filter(
            exists().where(
                (QuestionTag.question_id == Question.id) &
                (QuestionTag.tag_id == level_id)
            )
        )

    total = q.count()
    rows = (
        q.order_by(Question.id.asc())
         .offset((page - 1) * size)
         .limit(size)
         .all()
    )
    return total, rows

def get_my_questions(
    db: Session, page: int, size: int,
    keyword: str|None, qtype: str|None, difficulty: int|None, active_only: bool,
    subject_id: int | None = None, level_id: int | None = None
):
    q = (
        db.query(
            Question.id.label("question_id"),
            Question.type,
            Question.difficulty,
            Question.audit_status,
            Question.updated_at,
            Question.created_at,
            QuestionVersion.stem,
        )
        .join(QuestionVersion, QuestionVersion.id == Question.current_version_id, isouter=False)
        .filter(Question.is_active == True)
    )
    if keyword:
        q = q.filter(QuestionVersion.stem.ilike(f"%{keyword}%"))
    if qtype:
        q = q.filter(Question.type == qtype)
    if difficulty:
        q = q.filter(Question.difficulty == difficulty)
    if active_only:
        q = q.filter(Question.audit_status == "APPROVED")
    if subject_id:
        q = q.filter(
            exists().where(
                (QuestionTag.question_id == Question.id) &
                (QuestionTag.tag_id == subject_id)
            )
        )
    if level_id:
        q = q.filter(
            exists().where(
                (QuestionTag.question_id == Question.id) &
                (QuestionTag.tag_id == level_id)
            )
        )

    total = q.count()
    rows = (q
            .order_by(Question.id.asc())
            .offset((page-1)*size).limit(size).all())
    return {
        "total": total,
        "page": page,
        "items": [dict(r._mapping) for r in rows],
    }


def get_questions_brief(db: Session, id_list: list[int], user_id: int, is_admin: bool):
    """批量获取题目简要信息（题干/选项/解析）"""
    from app.core.exceptions import NotFoundException
    
    if not id_list:
        return []
    
    # 兼容字段名
    QV = QuestionVersion
    analysis_col = getattr(QV, "analysis", None) or getattr(QV, "explanation", None)
    options_col = getattr(QV, "options", None) or getattr(QV, "choices", None)
    
    cols = [Question.id.label("id"), QV.stem.label("stem")]
    if options_col is not None: 
        cols.append(options_col.label("options"))
    if analysis_col is not None: 
        cols.append(analysis_col.label("analysis"))
    
    q = (
        db.query(*cols)
          .outerjoin(QV, Question.current_version_id == QV.id)
          .filter(Question.id.in_(id_list))
    )
    
    # 🔒 非管理员只能查看自己创建的题目
    if not is_admin:
        q = q.filter(QV.created_by == user_id)
    
    rows = q.all()
    by_id = {r.id: r for r in rows}
    
    result = []
    for qid in id_list:
        r = by_id.get(qid)
        if r:
            result.append(r)
    
    return result


def get_question_detail(db: Session, qid: int, user_id: int, is_admin: bool):
    """获取单个题目详情"""
    from app.core.exceptions import NotFoundException, ForbiddenException
    
    QV = QuestionVersion
    analysis_col = getattr(QV, "analysis", None) or getattr(QV, "explanation", None)
    options_col = getattr(QV, "options", None) or getattr(QV, "choices", None)
    correct_answer_col = getattr(QV, "correct_answer", None) or getattr(QV, "answer", None)
    
    cols = [
        Question.id.label("id"), 
        Question.type.label("type"),
        Question.is_active.label("is_active"),
        QV.stem.label("stem")
    ]
    if options_col is not None: 
        cols.append(options_col.label("options"))
    if analysis_col is not None: 
        cols.append(analysis_col.label("analysis"))
    if correct_answer_col is not None: 
        cols.append(correct_answer_col.label("correct_answer"))
    
    q = (db.query(*cols)
            .outerjoin(QV, Question.current_version_id == QV.id)
            .filter(Question.id == qid))
    
    # 🔒 非管理员只能查看自己创建的题目
    if not is_admin:
        q = q.filter(QV.created_by == user_id)
    
    r = q.first()
    if not r:
        raise NotFoundException("题目不存在或无权限访问")
    
    return r


def get_question_owner_id(q: Question, db: Session) -> int | None:
    """获取题目的创建者ID"""
    # 1) Question.created_by 优先
    if hasattr(q, "created_by"):
        return getattr(q, "created_by")
    # 2) 回退到版本表的 created_by
    if hasattr(q, "current_version_id") and q.current_version_id:
        return db.query(QuestionVersion.created_by)\
                 .filter(QuestionVersion.id == q.current_version_id)\
                 .scalar()
    return None


def update_question(db: Session, qid: int, body, user_id: int, is_admin: bool):
    """更新题目信息"""
    from app.core.exceptions import NotFoundException, ForbiddenException
    
    q = db.query(Question).filter(Question.id == qid).first()
    if not q:
        raise NotFoundException("题目不存在")
    
    owner_id = get_question_owner_id(q, db)
    if not is_admin and (owner_id is not None) and (owner_id != user_id):
        raise ForbiddenException("无权限")
    
    # 获取题目版本（优先 current_version_id，其次按最新一条兜底）
    QV = QuestionVersion
    qv = None
    if hasattr(q, "current_version_id") and q.current_version_id:
        qv = db.query(QV).filter(QV.id == q.current_version_id).first()
    if not qv:
        qv = db.query(QV).filter(QV.question_id == q.id).order_by(QV.id.desc()).first()
    if not qv:
        raise NotFoundException("题目版本不存在")
    
    # 更新字段
    if body.stem is not None:
        qv.stem = body.stem.strip()
    
    if body.options is not None:
        val_list = _options_to_db_list(body.options)
        if val_list is not None:
            if hasattr(qv, "options"):
                qv.options = val_list
            elif hasattr(qv, "choices"):
                qv.choices = json.dumps(val_list, ensure_ascii=False)
    
    if body.analysis is not None:
        if hasattr(qv, "analysis"):
            qv.analysis = body.analysis
        elif hasattr(qv, "explanation"):
            qv.explanation = body.analysis
    
    if body.correct_answer is not None:
        ca = (body.correct_answer or "").strip().upper()[:8]
        if hasattr(qv, "correct_answer"):
            qv.correct_answer = ca
        elif hasattr(qv, "answer"):
            qv.answer = ca
    
    # is_active 应该更新到 Question 表
    if body.is_active is not None:
        q.is_active = bool(body.is_active)
        if hasattr(qv, "is_active"):
            qv.is_active = bool(body.is_active)
    
    # 更新题目类型
    if body.type is not None:
        allowed_types = ["SC", "MC", "FILL"]
        if body.type in allowed_types:
            q.type = body.type
        else:
            raise HTTPException(status_code=400, detail=f"题目类型必须是以下之一: {allowed_types}")
    
    # 保存时默认通过审核
    if hasattr(qv, "audit_status"):
        qv.audit_status = "APPROVED"
    elif hasattr(q, "audit_status"):
        q.audit_status = "APPROVED"
    
    db.commit()
    return {"ok": True}


def _options_to_db_list(val):
    """规范化前端传来的 options，返回 Python 列表[str]"""
    if val is None:
        return None
    try:
        if isinstance(val, str):
            parsed = json.loads(val)
            val = parsed
    except Exception:
        pass
    
    out = []
    if isinstance(val, list):
        for it in val:
            if isinstance(it, dict):
                out.append((it.get("text") or it.get("content") or "").strip())
            else:
                out.append(str(it))
        return out
    return [str(val)]


def list_tags(db: Session, type_filter: str | None = None):
    """获取标签列表"""
    q = db.query(Tag)
    if type_filter:
        q = q.filter(Tag.type == type_filter)
    rows = q.order_by(Tag.type.asc(), Tag.name.asc()).all()
    return rows


def get_question_tags(db: Session, qid: int, user_id: int, is_admin: bool):
    """获取题目的标签"""
    from app.core.exceptions import NotFoundException, ForbiddenException
    
    q = db.query(Question).filter(Question.id == qid).first()
    if not q:
        raise NotFoundException("题目不存在")
    
    owner_id = get_question_owner_id(q, db)
    if not is_admin and (owner_id is not None) and (owner_id != user_id):
        raise ForbiddenException("无权限访问此题目")
    
    rows = (
        db.query(QuestionTag.tag_id, Tag.type)
        .join(Tag, Tag.id == QuestionTag.tag_id)
        .filter(QuestionTag.question_id == qid)
        .all()
    )
    
    subject_id = next((tid for tid, tp in rows if tp == "SUBJECT"), None)
    level_id = next((tid for tid, tp in rows if tp == "LEVEL"), None)
    
    return {
        "subject_id": subject_id,
        "level_id": level_id,
        "tag_ids": [tid for tid, _ in rows],
    }


def set_question_tags(db: Session, qid: int, body, user_id: int, is_admin: bool):
    """设置题目的标签"""
    from app.core.exceptions import NotFoundException, ForbiddenException
    
    q = db.query(Question).filter(Question.id == qid).first()
    if not q:
        raise NotFoundException("题目不存在")
    
    owner_id = get_question_owner_id(q, db)
    if not is_admin and (owner_id is not None) and (owner_id != user_id):
        raise ForbiddenException("无权限")
    
    # SUBJECT 互斥
    if body.subject_id is not None:
        old_sids = [
            tid for (tid,) in db.query(QuestionTag.tag_id)
            .join(Tag, Tag.id == QuestionTag.tag_id)
            .filter(QuestionTag.question_id == qid, Tag.type == "SUBJECT").all()
        ]
        if old_sids:
            db.query(QuestionTag).filter(
                QuestionTag.question_id == qid,
                QuestionTag.tag_id.in_(old_sids)
            ).delete(synchronize_session=False)
        if body.subject_id:
            ok = db.query(Tag.id).filter(Tag.id == body.subject_id, Tag.type == "SUBJECT").first()
            if not ok:
                raise HTTPException(status_code=400, detail="subject_id 非法")
            exists = db.query(QuestionTag).filter(
                QuestionTag.question_id == qid, QuestionTag.tag_id == body.subject_id
            ).first()
            if not exists:
                db.add(QuestionTag(question_id=qid, tag_id=body.subject_id))
    
    # LEVEL 互斥
    if body.level_id is not None:
        old_lids = [
            tid for (tid,) in db.query(QuestionTag.tag_id)
            .join(Tag, Tag.id == QuestionTag.tag_id)
            .filter(QuestionTag.question_id == qid, Tag.type == "LEVEL").all()
        ]
        if old_lids:
            db.query(QuestionTag).filter(
                QuestionTag.question_id == qid,
                QuestionTag.tag_id.in_(old_lids)
            ).delete(synchronize_session=False)
        if body.level_id:
            ok = db.query(Tag.id).filter(Tag.id == body.level_id, Tag.type == "LEVEL").first()
            if not ok:
                raise HTTPException(status_code=400, detail="level_id 非法")
            exists = db.query(QuestionTag).filter(
                QuestionTag.question_id == qid, QuestionTag.tag_id == body.level_id
            ).first()
            if not exists:
                db.add(QuestionTag(question_id=qid, tag_id=body.level_id))
    
    # 可选批量增删
    if body.remove_ids:
        db.query(QuestionTag).filter(
            QuestionTag.question_id == qid,
            QuestionTag.tag_id.in_(body.remove_ids)
        ).delete(synchronize_session=False)
    if body.add_ids:
        for tid in body.add_ids:
            exists = db.query(QuestionTag).filter(
                QuestionTag.question_id == qid, QuestionTag.tag_id == tid
            ).first()
            if not exists:
                db.add(QuestionTag(question_id=qid, tag_id=tid))
    
    db.commit()
    return {"ok": True}